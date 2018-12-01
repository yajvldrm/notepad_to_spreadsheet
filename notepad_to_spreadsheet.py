import openpyxl
from openpyxl.utils.cell import column_index_from_string
"""
This example needs a exact format of a notepad

In our spreadsheet, the key is the 'Student Number' and the column to be filled is 'Quiz 2'
so the format of our text file must be like this:

format: key value
example: 15-9201 40

We must also know the column number of the 'Quiz 2'.
"""

txt_file_name = input("Enter the name fo the textfile: ")
spreadsheet = input("Enter the name of the workbook: ")
sheet_name = input("Enter the name of the sheet: ")
column_letter = input("Enter the letter ofthe column to be filled: ")
column_key = input("Enter the letter of the key column :")

txt_file = open(txt_file_name,'r')
wb = openpyxl.load_workbook(spreadsheet)
sheet = wb[sheet_name]
column_index = column_index_from_string(column_letter)
key_column_index = column_index_from_string(column_key)

grades = []
for line in txt_file:
    grades.append(line.split(' '))

for rowNum in range(2, sheet.max_row + 1):
    key = sheet.cell(row=rowNum, column = key_column_index).value
    for i in range(0, len(grades)):
        if key == grades[i][0]:
            sheet.cell(row=rowNum, column=column_index).value = grades[i][1]
            print("Student with Student number of " + key + " has the value of " +  str(grades[i][1]))

wb.save(spreadsheet)





